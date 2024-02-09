from flask import Flask , render_template , request , redirect , send_from_directory
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import re
import xlsxwriter
from xlsxwriter.utility import xl_col_to_name
app = Flask(__name__)

app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///ravisabha.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

class Ravisabha(db.Model):
    sno = db.Column(db.Integer,primary_key=True)
    name = db.Column(db.String(50),nullable = False)
    date1 = db.Column(db.DateTime,default=datetime.utcnow)
    
    def __repr__(self) -> str:
        return f"{self.sno} - {self.name} - {self.date1}"
    
#trigger at index.html go to index_data.html
@app.route('/display',methods=['GET','POST'])
def display_id(sno,names):
    return render_template('index_data.html',sno=sno,names=names)

#trigger at name.html go to name_data.html
@app.route('/name/display',methods=['GET','POST'])
def display_name(sno,names):
    return render_template('name_data.html',sno=sno,names=names)

#trigger at index_data.html go to index.html
@app.route('/submit',methods=['GET','POST'])
def submit_id_data():
    print(request.form['iddata'])
    if request.method == 'POST':
        userid = request.form['iddata']
        username = request.form['namedata']
        
        atten = Ravisabha(sno=userid,name=username)
        db.session.add(atten)
        db.session.commit()
    return render_template('index.html')

#trigger at name_data.html go to name.html
@app.route('/submitname/<int:s_no>/<n_ame>',methods=['GET','POST'])
def submit_name_data(s_no,n_ame):
    
    # print(request.form['iddata1'])
    if request.method == 'POST':
        # userid = request.form['iddata1']
        # username = request.form['namedata1']
        
        atten = Ravisabha(sno=s_no,name=n_ame)
        db.session.add(atten)
        db.session.commit()
    return render_template('name.html')

#trigger at index.html go to index_data.html
#trigger at name.html go to index.html
@app.route('/',methods=['GET','POST'])
def search_id():
    if request.method=='POST':
        userid1 = request.form['userid']
        userid = int(userid1)
        wb = openpyxl.load_workbook('Parents.xlsx')
        ws = wb.active
        # Find index of the ID in the excel file
        val=1234567
        for row in range(1, ws.max_row):
            for id in ws.iter_cols(1, 1):  # 1st columnn ie ID(Column) to start and end at 1st column 
                if id[row].value == userid: 
                    # print(f'id : {id[row].value} row : {row} userid : {userid , type(userid)}')
                    val = row
                    
        if val != 1234567:
            for index,col in enumerate(ws.iter_cols(1, ws.max_column)):
                    x = col[val].value
                    if type(x) == type(1):
                        s_no=col[val].value
                    elif type(x) == type('int'):
                        n_ame=col[val].value
                        
        else:
            return render_template('index.html',val=0)
        return render_template('index_data.html',s_no=s_no,n_ame=n_ame)
    return render_template('index.html',val=1)

#trigger at index.html go to name.html
#trigger at name.html go to name_data.html
@app.route('/name',methods=['GET','POST'])
def search_name():
    if request.method=='POST':
        username = request.form['username']
        username = username.lower()
        wb = openpyxl.load_workbook('Parents.xlsx')
        ws = wb.active
        # Find index of the Name in the excel file
        val=[]
        for row in range(1, ws.max_row):
            for col in ws.iter_cols(2, 2):  # 1st columnn ie ID(Column) to start and end at 1st column
                
                print(re.match(f"^{username}",col[row].value.lower()))
                if  re.match(f"^{username}",col[row].value.lower()): 
                    # print(f'name : {col[row].value} row : {row} username : {username , type(username)}')
                    val.append(row)
        s_no=[]
        n_ame=[]            
        if len(val)>0:
            print('2nd loop entered ===================================================')
            for index,col in enumerate(ws.iter_cols(1, ws.max_column)):
                for i in val:
                    print(col[i].value)
                    x = col[i].value
                    if type(x) == type(1):
                        s_no.append(col[i].value) #stores ID
                    elif type(x) == type('int'):
                        n_ame.append(col[i].value) #stores Name
        else:
            return render_template('name.html',val=0)
        return render_template('name_data.html',s_no=s_no,n_ame=n_ame)              
    return render_template('name.html',val=1)

#trigger at base.html go to table.html
@app.route('/data')
def data():
    attendence = Ravisabha.query.all()
    print(attendence)
    return render_template('table.html',attendence=attendence)

# @app.route('/uses')
# def use():
#     return 'Hello, uses!'

#trigger at table.html go to table.html
@app.route('/delete/<int:sno>')
def delete(sno):
    attendence = Ravisabha.query.filter_by(sno=sno).first()
    db.session.delete(attendence)
    db.session.commit()
    return redirect('/data')

#trigger at base.html go to download.html
@app.route('/download')
def download_excel():
    return send_from_directory( directory=app.static_folder , path='Parents.xlsx',)

@app.route('/write')
def write_excel():
     
    wb = openpyxl.load_workbook('./static/Parents.xlsx')
    ws = wb.active
    
    datalist = list(Ravisabha.query.all())
    # print('datalist = =',datalist)
    
    col1 = ws.max_column
    if col1 >= 4:
        col1 = xl_col_to_name(ws.max_column-1)
        col2 = xl_col_to_name(ws.max_column)
        print(col2,col1,'col2,col1 in IF')
    else:
        col1 = xl_col_to_name(ws.max_column)
        col2 = xl_col_to_name(ws.max_column+1)
        print(col2,col1,'col2,col1 in ELSE')
        
    
    # print(f"max column {ws.max_column} , col1 = {col1}")
    # col1 = xl_col_to_name(col1)
    
    for row in range(1,ws.max_row-1):
        
        ws[f'{col1}{row+1}'] = 'Absent'
        ws[f'{col1}{row+1}'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid") #red colour
            
        if row == 1:
            ws[f'{col1}{ws.max_row}'] = f'=COUNTIF({col1}{2}:{col1}{ws.max_row-1},"Present")'
            # ws[f'{col1}{row+1}'] = dat line at 187
            ws[f'{col2}{row}'] = "Individual Attendence"
    coldate = ''
    for i in range(0,len(datalist)):
        cols = str(datalist[i]) # get query data into string format
        print(cols)
        cols = cols.split(" - ") # data in fromat like in line no 19
        id = int(cols[0]) # 0 = ID in cols , 1 = Name , 2 = date
        dat = cols[2].split(" ") # get  date
        dat = dat[0].split("-")  # Got Date only in string format yyyy-mm-dd
        dat = dat[::-1]
        dat = str(dat[0]+'/'+dat[1]+'/'+dat[2]) # Date format dd/mm/yyyy
        coldate = dat
        if coldate!='':
            ws[f'{col1}{1}'] = coldate
        
        for row in range(1,ws.max_row-1):
            for id_column in ws.iter_cols(1, 1):  # 1st columnn ie ID(Column) to start and end at 1st column 
                # print(row,id_column[row].value , id)
                if id_column[row].value == id:
                    ws[f'{col1}{row+1}'] = 'Present'
                    ws[f'{col1}{row+1}'].fill = PatternFill(start_color="008000", end_color="008000", fill_type = "solid") # green color
                ws[f'{col2}{row+1}'] = f'=COUNTIF(C{row+1}:{col1}{row+1},"Present")'
                ws[f'{col2}{row+1}'].fill = PatternFill(start_color="0088AA", end_color="0088AA", fill_type = "solid") # Blue color
                
    wb.save('./static/Parents.xlsx')
    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True,port=8000)